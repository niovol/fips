import csv
from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Dict

import openpyxl
from openpyxl.styles import Alignment

from .logger import logger
from .models import PatentResult


class PatentStorage:
    """Handles storage operations for patent data."""

    def __init__(self, base_dir: Path):
        self.base_dir = base_dir
        patents_basename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        self.csv_path = base_dir / f"{patents_basename}.csv"
        self.patents_dir = base_dir / patents_basename
        self._setup_storage()

    def _setup_storage(self) -> None:
        """Initialize storage directories and files."""
        self.patents_dir.mkdir(parents=True, exist_ok=True)
        self._initialize_csv()
        logger.info(f"Storage initialized at {self.base_dir}")

    def _initialize_csv(self) -> None:
        """Create CSV file with headers."""
        headers = asdict(PatentResult("", "", "", "", "")).keys()
        with open(self.csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()

    def save_patent_to_csv(self, patent: PatentResult) -> None:
        """Save patent data to CSV."""
        with open(self.csv_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=asdict(patent).keys())
            writer.writerow(asdict(patent))

    def save_patent_details(self, patent_number: str, details: Dict) -> None:
        """Save detailed patent information to XLSX with specific field ordering."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Информация о патенте"

        # Priority fields to be placed at the top
        priority_fields = ["Ссылка", "Документ", "МПК"]

        # Start with priority fields
        current_row = 1
        for field in priority_fields:
            if field in details:
                name_cell = ws.cell(row=current_row, column=1, value=field)
                name_cell.alignment = Alignment(wrap_text=True, vertical="top")

                if field == "Ссылка":
                    cell = ws.cell(row=current_row, column=2, value=details[field])
                    cell.hyperlink = details[field]
                    cell.style = "Hyperlink"
                else:
                    cell = ws.cell(row=current_row, column=2, value=details[field])
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

                current_row += 1
                del details[field]

        # Add remaining fields
        for key, value in details.items():
            name_cell = ws.cell(row=current_row, column=1, value=key)
            name_cell.alignment = Alignment(wrap_text=True, vertical="top")

            if value:
                value_cell = ws.cell(row=current_row, column=2, value=value)
                value_cell.alignment = Alignment(wrap_text=True, vertical="top")

            current_row += 1

        self._adjust_column_widths(ws)
        self._adjust_row_heights(ws)

        file_path = self.patents_dir / f"{patent_number}.xlsx"
        wb.save(file_path)
        logger.info(f"Saved XLSX file: {file_path}")

    @staticmethod
    def _adjust_column_widths(worksheet) -> None:
        """Adjust column widths based on content."""
        for column in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            # Limit maximum width to 100 characters to prevent too wide columns
            adjusted_width = min(max_length + 2, 100)
            worksheet.column_dimensions[
                openpyxl.utils.get_column_letter(column[0].column)
            ].width = adjusted_width

    @staticmethod
    def _adjust_row_heights(worksheet) -> None:
        """Adjust row heights to fit multiline content."""
        DEFAULT_ROW_HEIGHT = 15  # Standard Excel row height
        MIN_HEIGHT_PADDING = 5  # Additional padding for each row
        CHAR_HEIGHT_FACTOR = 1.2  # Factor to account for character height variations

        for row in worksheet.rows:
            max_lines = 1

            for cell in row:
                if not cell.value:  # Skip empty cells
                    continue

                cell_text = str(cell.value)

                # 1. Count explicit line breaks
                explicit_lines = cell_text.count("\n") + 1

                # 2. Calculate wrapped lines based on column width
                column_width = worksheet.column_dimensions[
                    openpyxl.utils.get_column_letter(cell.column)
                ].width

                # Average chars that fit in column (assuming default font)
                chars_per_line = max(1, int(column_width * 1.8))

                # Calculate wrapped lines
                text_length = len(cell_text)
                wrapped_lines = -(-text_length // chars_per_line)  # Ceiling division

                # Take maximum of explicit and wrapped lines
                total_lines = max(explicit_lines, wrapped_lines)

                # Update max_lines if this cell needs more
                max_lines = max(max_lines, total_lines)

            # Calculate final row height with padding
            if max_lines > 1:
                # Base height calculation
                base_height = max_lines * DEFAULT_ROW_HEIGHT

                # Add padding and adjust by factor
                final_height = (base_height * CHAR_HEIGHT_FACTOR) + MIN_HEIGHT_PADDING

                # Set the row height
                worksheet.row_dimensions[row[0].row].height = final_height
